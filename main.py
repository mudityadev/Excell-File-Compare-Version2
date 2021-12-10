import pandas as pd 
from openpyxl import load_workbook
import sys

excel_file_1 = sys.argv[1]
excel_file_2 = sys.argv[2]

df1 = pd.read_excel(excel_file_1)
df2 = pd.read_excel(excel_file_2)




# print(df1['File Name'].isin(df2['File Name']))

filtered_df_1 = df1.loc[(df1['FileName'].isin(df2['FileName']))]
# print(filtered_df_1)
# filtered_df_1.drop(columns=[' '])
filtered_df_2 = df1.loc[~(df1['FileName'].isin(df2['FileName']))]
# print(filtered_df_2)
# filtered_df_1.style.hide_index()
filtered_df_2.to_excel("results/output.xlsx", index=False)
output = pd.read_excel('results/output.xlsx')

print(output)
# print(output.head())

print( "[INFO] Size of remaining rows and column = "+ str(output.shape))
# # wb = load_workbook('output.xlsx')
# # ws = wb.active
# # ws['A1'].value = "Start"
# # wb.save('output.xlsx')
